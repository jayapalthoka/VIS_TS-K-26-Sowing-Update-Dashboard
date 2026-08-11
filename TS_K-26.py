import streamlit as st
import pandas as pd
import plotly.express as px
import io
import os
from datetime import datetime
from zoneinfo import ZoneInfo

import gspread
from google.oauth2.service_account import Credentials
from streamlit_autorefresh import st_autorefresh

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ==========================================================
# PAGE CONFIG
# ==========================================================

st.set_page_config(
    page_title="VIS TS Kharif-26 Dashboard",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Refresh Google Sheet data every 60 seconds
st_autorefresh(interval=60000, key="dashboard_refresh")


# ==========================================================
# CUSTOM CSS
# ==========================================================

st.markdown(
    """
    <style>

    /* ======================================================
       GLOBAL DASHBOARD THEME
       ====================================================== */

    .main {
        background: #F7F9F7;
    }

    .block-container {
        padding-top: 0.35rem;
        padding-bottom: 0.65rem;
        padding-left: 1.5rem;
        padding-right: 1.5rem;
    }

    /* ======================================================
       KPI CARDS
       ====================================================== */

    .kpi-card {
        background: #FFFFFF;
        border: 1px solid #E0E7E0;
        border-radius: 14px;
        padding: 14px 16px 12px 16px;
        margin: 0 0 8px 0;
        box-shadow: 0 2px 8px rgba(30, 60, 30, .06);
    }

    .kpi-title {
        font-size: 1.15rem;
        font-weight: 700;
        color: #243024;
        margin-bottom: 2px;
    }

    .kpi-percent {
        font-size: 2rem;
        font-weight: 800;
        color: #2E7D32;
        line-height: 1.1;
        margin: 4px 0 7px 0;
    }

    .kpi-progress {
        height: 9px;
        width: 100%;
        background: #E8F1E8;
        border-radius: 10px;
        overflow: hidden;
        margin-bottom: 9px;
    }

    .kpi-progress-fill {
        height: 100%;
        background: #2E7D32;
        border-radius: 10px;
    }

    .kpi-row {
        display: flex;
        justify-content: space-between;
        align-items: center;
        font-size: .88rem;
        color: #777;
        padding: 3px 0;
    }

    .kpi-row b {
        color: #263238;
    }

    .kpi-row.achieved b {
        color: #2E7D32;
    }

    .kpi-row.pending b {
        color: #D32F2F;
    }

    /* ======================================================
       SIDEBAR
       ====================================================== */

    section[data-testid="stSidebar"] {
        background: #FFFFFF;
        border-right: 1px solid #E2E8E2;
    }

    /* ======================================================
       TAB BAR
       Different accent for each dashboard section
       ====================================================== */

    div[data-baseweb="tab-list"] {
        gap: 4px;
        border-bottom: 1px solid #DDE5DD;
        padding-bottom: 1px;
    }

    div[data-baseweb="tab-list"] button {
        border-radius: 9px 9px 0 0;
        padding: 6px 12px;
        font-weight: 600;
        transition: all .15s ease;
    }

    /* Overview - Green */
    div[data-baseweb="tab-list"] button:nth-child(1) {
        color: #2E7D32;
    }

    div[data-baseweb="tab-list"] button:nth-child(1)[aria-selected="true"] {
        background: #E8F5E9;
        color: #1B5E20;
        border-bottom: 3px solid #2E7D32;
    }

    /* Study & Cluster - Blue */
    div[data-baseweb="tab-list"] button:nth-child(2) {
        color: #1565C0;
    }

    div[data-baseweb="tab-list"] button:nth-child(2)[aria-selected="true"] {
        background: #E3F2FD;
        color: #0D47A1;
        border-bottom: 3px solid #1565C0;
    }

    /* Field Officer - Purple */
    div[data-baseweb="tab-list"] button:nth-child(3) {
        color: #6A1B9A;
    }

    div[data-baseweb="tab-list"] button:nth-child(3)[aria-selected="true"] {
        background: #F3E5F5;
        color: #4A148C;
        border-bottom: 3px solid #6A1B9A;
    }

    /* Timeline - Orange */
    div[data-baseweb="tab-list"] button:nth-child(4) {
        color: #EF6C00;
    }

    div[data-baseweb="tab-list"] button:nth-child(4)[aria-selected="true"] {
        background: #FFF3E0;
        color: #E65100;
        border-bottom: 3px solid #EF6C00;
    }

    /* Data - Teal */
    div[data-baseweb="tab-list"] button:nth-child(5) {
        color: #00796B;
    }

    div[data-baseweb="tab-list"] button:nth-child(5)[aria-selected="true"] {
        background: #E0F2F1;
        color: #004D40;
        border-bottom: 3px solid #00796B;
    }

    /* ======================================================
       STUDY & CLUSTER SECTION ACCENTS
       ====================================================== */

    .section-blue {
        background: #E3F2FD;
        border-left: 5px solid #1565C0;
        border-radius: 8px;
        padding: 8px 14px;
        margin: 10px 0 12px 0;
        color: #0D47A1;
        font-weight: 700;
    }

    .section-green {
        background: #E8F5E9;
        border-left: 5px solid #2E7D32;
        border-radius: 8px;
        padding: 8px 14px;
        margin: 10px 0 12px 0;
        color: #1B5E20;
        font-weight: 700;
    }

    .section-purple {
        background: #F3E5F5;
        border-left: 5px solid #6A1B9A;
        border-radius: 8px;
        padding: 8px 14px;
        margin: 10px 0 12px 0;
        color: #4A148C;
        font-weight: 700;
    }

    .section-orange {
        background: #FFF3E0;
        border-left: 5px solid #EF6C00;
        border-radius: 8px;
        padding: 8px 14px;
        margin: 10px 0 12px 0;
        color: #E65100;
        font-weight: 700;
    }

    .section-teal {
        background: #E0F2F1;
        border-left: 5px solid #00796B;
        border-radius: 8px;
        padding: 8px 14px;
        margin: 10px 0 12px 0;
        color: #004D40;
        font-weight: 700;
    }

    /* ======================================================
       DATAFRAME / TABLES
       ====================================================== */

    div[data-testid="stDataFrame"] {
        border: 1px solid #DDE5DD;
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 1px 5px rgba(0,0,0,.04);
    }

    /* ======================================================
       SINGLE OVERALL DOWNLOAD BUTTON
       ====================================================== */

    div[data-testid="stDownloadButton"] button[kind="secondary"] {
        min-height: 44px;
        font-size: 15px;
        font-weight: 700;
        border-radius: 10px;
        border: 1px solid #2E7D32;
    }

    /* ======================================================
       BUTTONS
       ====================================================== */

    div.stButton > button,
    div[data-testid="stDownloadButton"] button {
        border-radius: 8px;
        font-weight: 600;
    }

    div[data-testid="stDownloadButton"] {
        margin-bottom: 10px;
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ==========================================================
# HEADER
# ==========================================================

india_time = datetime.now(ZoneInfo("Asia/Kolkata"))

st.markdown(
    "<h1 style='margin:0 0 2px 0; padding:0; font-size:2.35rem;'>🌾 TS Kharif-26 Monitoring Dashboard</h1>",
    unsafe_allow_html=True
)

st.caption(
    f"Last Updated : {india_time.strftime('%d-%b-%Y %I:%M %p')}"
)


# ==========================================================
# GOOGLE SHEETS CONFIGURATION
# ==========================================================

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly"
]

SPREADSHEET_ID = "1XNzggS3DKn8ucffUx1QKBlrDXVLigLg3psR2WHdyR9I"
WORKSHEET_NAME = "Overall Farmer List - Kharif-26"


# ==========================================================
# LOAD DATA
# ==========================================================

@st.cache_data(ttl=60, show_spinner=False)
def load_data():

    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=SCOPES
    )

    client = gspread.authorize(creds)

    worksheet = (
        client
        .open_by_key(SPREADSHEET_ID)
        .worksheet(WORKSHEET_NAME)
    )

    df = pd.DataFrame(worksheet.get_all_records())

    # Clean column names
    df.columns = (
        df.columns
        .str.strip()
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
    )

    # Numeric columns
    numeric_columns = [
        "K/R-25 Hectares",
        "Kharif-26 Hectares",
        "Kharif-26 Acreage"
    ]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            ).fillna(0)

    # Date column
    date_col = "Kharif-26 DSR/Nursery Sowing Date"

    if date_col in df.columns:
        df[date_col] = pd.to_datetime(
            df[date_col],
            errors="coerce",
            dayfirst=True
        )

    # Active / Inactive status
    status_col = "Kharif-26 Plot Status (Active / Inactive )"

    if status_col in df.columns:
        df[status_col] = (
            df[status_col]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

    return df


try:
    with st.spinner("Loading Dashboard..."):
        df = load_data()

except Exception as e:
    st.error(
        f"Unable to load Google Sheet.\n\n{e}"
    )
    st.stop()


# ==========================================================
# SIDEBAR FILTERS
# ==========================================================

st.sidebar.title("Dashboard Filters")

study_filter = st.sidebar.multiselect(
    "Study",
    sorted(df["Study2"].dropna().unique())
)

cluster_filter = st.sidebar.multiselect(
    "Cluster",
    sorted(df["Cluster"].dropna().unique())
)

tahsil_filter = st.sidebar.multiselect(
    "Tahsil",
    sorted(df["Tahsil"].dropna().unique())
)

fo_filter = st.sidebar.multiselect(
    "Field Officer",
    sorted(df["Field Officer"].dropna().unique())
)

cp_filter = st.sidebar.multiselect(
    "Cultivation Practice",
    sorted(
        df["Kharif-26 Cultivation Practice"]
        .dropna()
        .astype(str)
        .unique()
    )
)


# ==========================================================
# APPLY FILTERS
# ==========================================================

filtered_df = df.copy()

if study_filter:
    filtered_df = filtered_df[
        filtered_df["Study2"].isin(study_filter)
    ]

if cluster_filter:
    filtered_df = filtered_df[
        filtered_df["Cluster"].isin(cluster_filter)
    ]

if tahsil_filter:
    filtered_df = filtered_df[
        filtered_df["Tahsil"].isin(tahsil_filter)
    ]

if fo_filter:
    filtered_df = filtered_df[
        filtered_df["Field Officer"].isin(fo_filter)
    ]

if cp_filter:
    filtered_df = filtered_df[
        filtered_df["Kharif-26 Cultivation Practice"].isin(cp_filter)
    ]


# ==========================================================
# ACTIVE DATA
# ==========================================================

STATUS_COL = "Kharif-26 Plot Status (Active / Inactive )"

filtered_df[STATUS_COL] = (
    filtered_df[STATUS_COL]
    .fillna("")
    .astype(str)
    .str.strip()
    .str.upper()
)

active_df = filtered_df[
    filtered_df[STATUS_COL].eq("ACTIVE")
].copy()


# ==========================================================
# KPI CALCULATIONS
# ==========================================================

target_farmers = filtered_df["Farmer Code"].nunique()

target_plots = filtered_df["Plot Codes"].nunique()

target_hectares = round(
    filtered_df["K/R-25 Hectares"].sum(),
    2
)

achieved_plots = active_df["Plot Codes"].nunique()

achieved_hectares = round(
    active_df["Kharif-26 Hectares"].sum(),
    2
)

pending_plots = max(
    target_plots - achieved_plots,
    0
)

pending_hectares = round(
    target_hectares - achieved_hectares,
    2
)

achievement_pct = round(
    achieved_hectares / target_hectares * 100
    if target_hectares > 0 else 0,
    2
)

pending_pct = round(
    100 - achievement_pct,
    2
)


# ==========================================================
# SUMMARY FUNCTION
# ==========================================================

@st.cache_data(show_spinner=False)
def _cultivation_summary_cached(
    filtered_data,
    active_data,
    group_column
):

    target = (
        filtered_data
        .groupby(group_column)
        .agg(
            Target_Farmers=("Farmer Code", "nunique"),
            Target_Plots=("Plot Codes", "nunique"),
            Target_Hectares=("K/R-25 Hectares", "sum")
        )
    )

    achieved = (
        active_data
        .groupby(group_column)
        .agg(
            Achieved_Plots=("Plot Codes", "nunique"),
            Achieved_Hectares=("Kharif-26 Hectares", "sum")
        )
    )

    summary = (
        target
        .join(achieved, how="left")
        .fillna(0)
    )

    summary["Pending_Plots"] = (
        summary["Target_Plots"]
        - summary["Achieved_Plots"]
    )

    summary["Pending_Hectares"] = (
        summary["Target_Hectares"]
        - summary["Achieved_Hectares"]
    )

    summary["Achieved %"] = (
        summary["Achieved_Hectares"]
        / summary["Target_Hectares"]
        * 100
    ).fillna(0).round(2)

    summary["Pending %"] = (
        100 - summary["Achieved %"]
    ).round(2)

    return summary.reset_index()


def cultivation_summary(group_column):
    return _cultivation_summary_cached(
        filtered_df,
        active_df,
        group_column
    )


# ==========================================================
# CULTIVATION PRACTICE SUMMARY
# ==========================================================

@st.cache_data(show_spinner=False)
def _cultivation_practice_summary_cached(
    filtered_data,
    active_data,
    group_column
):

    target = (
        filtered_data
        .groupby(group_column)
        .agg(
            Target_Ha=("K/R-25 Hectares", "sum")
        )
    )

    achieved = (
        active_data
        .groupby(group_column)
        .agg(
            Achieved_Ha=("Kharif-26 Hectares", "sum")
        )
    )

    summary = (
        target
        .join(achieved, how="left")
        .fillna(0)
    )

    practice_col = "Kharif-26 Cultivation Practice"
    hectare_col = "Kharif-26 Hectares"

    dry = (
        active_data[
            active_data[practice_col]
            .astype(str)
            .str.strip()
            .eq("Dry DSR")
        ]
        .groupby(group_column)[hectare_col]
        .sum()
    )

    wet = (
        active_data[
            active_data[practice_col]
            .astype(str)
            .str.strip()
            .eq("WET DSR")
        ]
        .groupby(group_column)[hectare_col]
        .sum()
    )

    tpr = (
        active_data[
            active_data[practice_col]
            .astype(str)
            .str.strip()
            .eq("Transplanting+AWD")
        ]
        .groupby(group_column)[hectare_col]
        .sum()
    )

    summary["Dry DSR (Ha)"] = dry
    summary["WET DSR (Ha)"] = wet
    summary["TPR+AWD (Ha)"] = tpr

    summary = summary.fillna(0)

    summary["Achieved %"] = (
        summary["Achieved_Ha"]
        / summary["Target_Ha"]
        * 100
    ).fillna(0).round(2)

    summary["Dry DSR %"] = (
        summary["Dry DSR (Ha)"]
        / summary["Achieved_Ha"]
        * 100
    ).fillna(0).round(2)

    summary["WET DSR %"] = (
        summary["WET DSR (Ha)"]
        / summary["Achieved_Ha"]
        * 100
    ).fillna(0).round(2)

    summary["TPR+AWD %"] = (
        summary["TPR+AWD (Ha)"]
        / summary["Achieved_Ha"]
        * 100
    ).fillna(0).round(2)

    return summary.reset_index()


def cultivation_practice_summary(group_column):
    return _cultivation_practice_summary_cached(
        filtered_df,
        active_df,
        group_column
    )


# ==========================================================
# SIMPLE EXCEL EXPORT
# ==========================================================

@st.cache_data(show_spinner=False)
def convert_excel(data):

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        data.to_excel(
            writer,
            sheet_name="Filtered Data",
            index=False
        )

    return output.getvalue()


# ==========================================================
# EXCEL STYLING
# ==========================================================

def _style_excel_sheet(
    ws,
    freeze_panes=None,
    tab_color="2E7D32"
):

    thin = Side(
        style="thin",
        color="D9E2D9"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # Do not freeze any row/column in the downloaded workbook.
    ws.freeze_panes = None
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

    for row in ws.iter_rows():

        for cell in row:

            if cell.value is None:
                continue

            cell.border = border

            cell.font = Font(
                name="Calibri",
                size=10,
                color="263238"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # Header
    for cell in ws[1]:

        cell.fill = PatternFill(
            "solid",
            fgColor="12355B"
        )

        cell.font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="FFFFFF"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    # Alternating rows
    for r in range(2, ws.max_row + 1):

        if r % 2 == 0:

            for c in range(1, ws.max_column + 1):

                ws.cell(r, c).fill = PatternFill(
                    "solid",
                    fgColor="F7FBF7"
                )

    # Number formats
    for c in range(1, ws.max_column + 1):

        header = ws.cell(1, c).value

        for r in range(2, ws.max_row + 1):

            cell = ws.cell(r, c)

            if header in (
                "Achieved %",
                "Pending %",
                "Dry DSR %",
                "WET DSR %",
                "TPR+AWD %"
            ):
                cell.number_format = '0.00"%"'

            elif (
                isinstance(header, str)
                and (
                    "Hectares" in header
                    or "Ha)" in header
                    or header.endswith("_Ha")
                )
            ):
                cell.number_format = "#,##0.00"

    # Smart column widths
    for col_cells in ws.columns:

        col_letter = get_column_letter(
            col_cells[0].column
        )

        max_length = 0

        for cell in col_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[col_letter].width = min(
            max(max_length + 3, 12),
            30
        )


def _style_study_cluster_sheet(ws):

    thin = Side(
        style="thin",
        color="D9E2D9"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    section_names = {
        "STUDY-WISE PROGRESS",
        "STUDY-WISE CULTIVATION",
        "CLUSTER-WISE PROGRESS",
        "CLUSTER-WISE CULTIVATION"
    }

    ws.sheet_view.showGridLines = False
    # Do not freeze any row/column in the downloaded workbook.
    ws.freeze_panes = None
    ws.sheet_properties.tabColor = "2E7D32"

    section_rows = []

    for r in range(1, ws.max_row + 1):

        value = ws.cell(r, 1).value

        if value in section_names:

            section_rows.append(r)

            if ws.max_column > 1:

                ws.merge_cells(
                    start_row=r,
                    start_column=1,
                    end_row=r,
                    end_column=ws.max_column
                )

            cell = ws.cell(r, 1)

            cell.fill = PatternFill(
                "solid",
                fgColor="1B5E20"
            )

            cell.font = Font(
                name="Calibri",
                size=12,
                bold=True,
                color="FFFFFF"
            )

            cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

            ws.row_dimensions[r].height = 26

    for r in range(1, ws.max_row + 1):

        if r in section_rows:
            continue

        for c in range(1, ws.max_column + 1):

            cell = ws.cell(r, c)

            if cell.value is None:
                continue

            cell.border = border

            cell.font = Font(
                name="Calibri",
                size=10,
                color="263238"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            if r % 2 == 0:

                cell.fill = PatternFill(
                    "solid",
                    fgColor="F7FBF7"
                )

    # Table headers are immediately after section titles
    for section_row in section_rows:

        header_row = section_row + 1

        for c in range(
            1,
            ws.max_column + 1
        ):

            cell = ws.cell(
                header_row,
                c
            )

            if cell.value is None:
                continue

            cell.fill = PatternFill(
                "solid",
                fgColor="2E7D32"
            )

            cell.font = Font(
                name="Calibri",
                size=10,
                bold=True,
                color="FFFFFF"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = border

        ws.row_dimensions[
            header_row
        ].height = 34

    for col in range(
        1,
        ws.max_column + 1
    ):

        ws.column_dimensions[
            get_column_letter(col)
        ].width = 18

    if ws.max_column >= 1:
        ws.column_dimensions["A"].width = 26


# ==========================================================
# OVERALL EXCEL CREATION
# ==========================================================

@st.cache_data(show_spinner=False)
def convert_dashboard_excel(
    study_summary,
    cluster_summary,
    study_cp,
    cluster_cp,
    fo_summary,
    timeline_data,
    raw_data
):

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        sheet_name = "Study & Cluster"

        start_row = 0

        # Study progress
        pd.DataFrame(
            [["STUDY-WISE PROGRESS"]]
        ).to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False,
            header=False
        )

        start_row += 1

        study_summary.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False
        )

        start_row += len(study_summary) + 3

        # Study cultivation
        pd.DataFrame(
            [["STUDY-WISE CULTIVATION"]]
        ).to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False,
            header=False
        )

        start_row += 1

        study_cp.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False
        )

        start_row += len(study_cp) + 3

        # Cluster progress
        pd.DataFrame(
            [["CLUSTER-WISE PROGRESS"]]
        ).to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False,
            header=False
        )

        start_row += 1

        cluster_summary.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False
        )

        start_row += len(cluster_summary) + 3

        # Cluster cultivation
        pd.DataFrame(
            [["CLUSTER-WISE CULTIVATION"]]
        ).to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False,
            header=False
        )

        start_row += 1

        cluster_cp.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False
        )

        # Other sheets
        fo_summary.to_excel(
            writer,
            sheet_name="Field Officer",
            index=False
        )

        timeline_data.to_excel(
            writer,
            sheet_name="Timeline",
            index=False
        )

        raw_data.to_excel(
            writer,
            sheet_name="Raw Data",
            index=False
        )

    # Open workbook for formatting
    wb = load_workbook(
        io.BytesIO(output.getvalue())
    )

    _style_study_cluster_sheet(
        wb["Study & Cluster"]
    )

    _style_excel_sheet(
        wb["Field Officer"],
        freeze_panes=None,
        tab_color="1565C0"
    )

    _style_excel_sheet(
        wb["Timeline"],
        freeze_panes=None,
        tab_color="EF6C00"
    )

    _style_excel_sheet(
        wb["Raw Data"],
        freeze_panes=None,
        tab_color="6A1B9A"
    )

    # Metadata
    wb.properties.title = (
        "TS Kharif-26 Overall Dashboard Report"
    )

    wb.properties.subject = (
        "TS Kharif-26 Sowing Progress"
    )

    wb.properties.creator = "VIS Dashboard"

    final_output = io.BytesIO()

    wb.save(final_output)

    return final_output.getvalue()


# ==========================================================
# MASTER DOWNLOAD DATA
# ==========================================================

master_study_summary = cultivation_summary("Study2")
master_cluster_summary = cultivation_summary("Cluster")

master_study_cp = cultivation_practice_summary("Study2")
master_cluster_cp = cultivation_practice_summary("Cluster")

master_fo_summary = cultivation_summary("Field Officer")


# Timeline master data
date_col = "Kharif-26 DSR/Nursery Sowing Date"

master_timeline_df = filtered_df.copy()

master_timeline_df = master_timeline_df[
    master_timeline_df[date_col].notna()
]

if not master_timeline_df.empty:

    master_daily = (
        master_timeline_df
        .groupby(date_col)
        .agg(
            Plots=("Plot Codes", "nunique"),
            Hectares=("Kharif-26 Hectares", "sum")
        )
        .reset_index()
        .sort_values(date_col)
    )

else:

    master_daily = pd.DataFrame(
        columns=[
            date_col,
            "Plots",
            "Hectares"
        ]
    )


# ==========================================================
# SINGLE OVERALL EXCEL DOWNLOAD
# ==========================================================

# Build the complete workbook once and cache it.
# Users only need to click ONE button to download the full report.
master_excel = convert_dashboard_excel(
    study_summary=master_study_summary,
    cluster_summary=master_cluster_summary,
    study_cp=master_study_cp,
    cluster_cp=master_cluster_cp,
    fo_summary=master_fo_summary,
    timeline_data=master_daily,
    raw_data=filtered_df
)

# ==========================================================
# COMPACT OVERALL DOWNLOAD BUTTON
# ==========================================================

# Keep the download control on the right side so it does not
# consume a full-width row or create unnecessary blank space.

_, download_col = st.columns(
    [8.5, 2.5],
    gap="small"
)

with download_col:

    st.download_button(
        label="📥 Download Overall Data",
        data=master_excel,
        file_name="TS_Kharif_26_Overall_Master.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key="overall_master_download",
        use_container_width=True
    )

# ==========================================================
# TABS
# ==========================================================

overview_tab, study_tab, fo_tab, timeline_tab, data_tab = st.tabs(
    [
        "📊 Overview",
        "🧭 Study & Cluster",
        "👨‍🌾 Field Officer",
        "📅 Timeline",
        "📋 Data"
    ]
)


# ==========================================================
# REUSABLE KPI CARD
# ==========================================================

def summary_card(
    title,
    icon,
    target,
    achieved,
    unit="",
    decimals=0
):

    pending = max(
        target - achieved,
        0
    )

    percent = (
        achieved / target * 100
        if target > 0 else 0
    )

    percent = min(
        percent,
        100
    )

    if decimals == 0:

        target_text = f"{target:,.0f}"
        achieved_text = f"{achieved:,.0f}"
        pending_text = f"{pending:,.0f}"

    else:

        target_text = (
            f"{target:,.2f}{unit}"
        )

        achieved_text = (
            f"{achieved:,.2f}{unit}"
        )

        pending_text = (
            f"{pending:,.2f}{unit}"
        )

    kpi_html = f"""
    <div class="kpi-card">
        <div class="kpi-title">{icon} {title}</div>
        <div class="kpi-percent">{percent:.1f}%</div>
        <div class="kpi-progress">
            <div class="kpi-progress-fill" style="width:{percent:.1f}%"></div>
        </div>
        <div class="kpi-row">
            <span>🎯 Target</span>
            <b>{target_text}</b>
        </div>
        <div class="kpi-row achieved">
            <span>✅ Achieved</span>
            <b>{achieved_text}</b>
        </div>
        <div class="kpi-row pending">
            <span>⏳ Pending</span>
            <b>{pending_text}</b>
        </div>
    </div>
    """

    # st.html prevents Streamlit from displaying the HTML source as text.
    if hasattr(st, "html"):
        st.html(kpi_html)
    else:
        st.markdown(
            kpi_html,
            unsafe_allow_html=True
        )


# ==========================================================
# OVERVIEW TAB
# ==========================================================

with overview_tab:

    st.subheader(
        "📊 Dashboard Overview"
    )

    # KPI cards
    c1, c2, c3 = st.columns(3)

    achieved_farmers = (
        active_df["Farmer Code"].nunique()
    )

    with c1:

        summary_card(
            title="Farmers",
            icon="👨‍🌾",
            target=target_farmers,
            achieved=achieved_farmers
        )

    with c2:

        summary_card(
            title="Plots",
            icon="📍",
            target=target_plots,
            achieved=achieved_plots
        )

    with c3:

        summary_card(
            title="Area (Ha)",
            icon="🌱",
            target=target_hectares,
            achieved=achieved_hectares,
            unit=" Ha",
            decimals=2
        )

    st.divider()


    # ======================================================
    # TWO MEDIUM-SIZE DONUT CHARTS
    # ======================================================

    chart_left, chart_right = st.columns(
        2,
        gap="large"
    )


    # ======================================================
    # LEFT: TARGET VS ACHIEVED HECTARES
    # ======================================================

    with chart_left:

        st.markdown(
            "### 🎯 Target vs Achieved Hectares"
        )

        target_chart_df = pd.DataFrame(
            {
                "Status": [
                    "Achieved",
                    "Pending"
                ],
                "Hectares": [
                    achieved_hectares,
                    pending_hectares
                ]
            }
        )

        fig1 = px.pie(
            target_chart_df,
            names="Status",
            values="Hectares",
            hole=0.62,
            color="Status",
            color_discrete_map={
                "Achieved": "#2E7D32",
                "Pending": "#EF6C00"
            }
        )

        fig1.update_traces(
            textposition="inside",
            texttemplate=(
                "<b>%{label}</b><br>"
                "%{value:,.2f} Ha<br>"
                "%{percent}"
            ),
            textfont=dict(
                size=13,
                color="white"
            ),
            hovertemplate=(
                "<b>%{label}</b><br>"
                "%{value:,.2f} Ha<br>"
                "%{percent}<extra></extra>"
            )
        )

        fig1.add_annotation(
            x=0.5,
            y=0.5,
            xref="paper",
            yref="paper",
            text=(
                "<b>Total Target</b><br>"
                f"<b style='font-size:22px'>"
                f"{target_hectares:,.2f} Ha"
                f"</b>"
            ),
            showarrow=False,
            font=dict(
                size=14,
                color="#263238"
            ),
            align="center"
        )

        fig1.update_layout(
            height=360,
            margin=dict(
                l=10,
                r=10,
                t=10,
                b=45
            ),
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.10,
                xanchor="center",
                x=0.5,
                font=dict(size=12)
            ),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)"
        )

        st.plotly_chart(
            fig1,
            use_container_width=True,
            config={
                "displayModeBar": False
            }
        )


    # ======================================================
    # RIGHT: CULTIVATION PRACTICE — % OF TOTAL TARGET
    # ======================================================

    with chart_right:

        st.markdown(
            "### 🌾 Cultivation Practice — Hectares"
        )

        practice_col = (
            "Kharif-26 Cultivation Practice"
        )

        hectare_col = (
            "Kharif-26 Hectares"
        )

        # IMPORTANT:
        # Practice hectares are calculated from ACTIVE plots only.
        # Pending is always Target - Achieved.
        # All percentages are calculated against TOTAL TARGET.

        dry_dsr_ha = round(
            active_df.loc[
                active_df[practice_col]
                .astype(str)
                .str.strip()
                .eq("Dry DSR"),
                hectare_col
            ].sum(),
            2
        )

        wet_dsr_ha = round(
            active_df.loc[
                active_df[practice_col]
                .astype(str)
                .str.strip()
                .eq("WET DSR"),
                hectare_col
            ].sum(),
            2
        )

        tpr_awd_ha = round(
            active_df.loc[
                active_df[practice_col]
                .astype(str)
                .str.strip()
                .eq("Transplanting+AWD"),
                hectare_col
            ].sum(),
            2
        )

        cultivation_pending_ha = round(
            target_hectares
            - achieved_hectares,
            2
        )

        cultivation_chart_df = pd.DataFrame(
            {
                "Cultivation Practice": [
                    "Dry DSR",
                    "WET DSR",
                    "Transplanting + AWD",
                    "Pending"
                ],
                "Hectares": [
                    dry_dsr_ha,
                    wet_dsr_ha,
                    tpr_awd_ha,
                    cultivation_pending_ha
                ]
            }
        )

        cultivation_chart_df[
            "Percentage"
        ] = (
            cultivation_chart_df["Hectares"]
            / target_hectares
            * 100
            if target_hectares > 0
            else 0
        )

        # Keep the cultivation practice order fixed:
        # 1. Dry DSR
        # 2. WET DSR
        # 3. Transplanting + AWD
        # 4. Pending
        cultivation_order = [
            "Dry DSR",
            "WET DSR",
            "Transplanting + AWD",
            "Pending"
        ]

        cultivation_chart_df[
            "Cultivation Practice"
        ] = pd.Categorical(
            cultivation_chart_df["Cultivation Practice"],
            categories=cultivation_order,
            ordered=True
        )

        cultivation_chart_df = (
            cultivation_chart_df
            .sort_values("Cultivation Practice")
            .reset_index(drop=True)
        )

        fig2 = px.pie(
            cultivation_chart_df,
            names="Cultivation Practice",
            values="Hectares",
            hole=0.62,
            color="Cultivation Practice",
            category_orders={
                "Cultivation Practice": cultivation_order
            },
            color_discrete_map={
                "Dry DSR": "#1B5E20",
                "WET DSR": "#66BB6A",
                "Transplanting + AWD": "#EF9A9A",
                "Pending": "#D32F2F"
            }
        )

        # Prevent Plotly from re-sorting slices by value.
        fig2.update_traces(sort=False)

        fig2.update_traces(
            textposition="inside",
            texttemplate=(
                "<b>%{label}</b><br>"
                "%{value:,.2f} Ha<br>"
                "%{percent}"
            ),
            textfont=dict(
                size=12,
                color="white"
            ),
            hovertemplate=(
                "<b>%{label}</b><br>"
                "%{value:,.2f} Ha<br>"
                "%{percent} of Target"
                "<extra></extra>"
            )
        )

        fig2.add_annotation(
            x=0.5,
            y=0.5,
            xref="paper",
            yref="paper",
            text=(
                "<b>Total Target</b><br>"
                f"<b style='font-size:22px'>"
                f"{target_hectares:,.2f} Ha"
                f"</b>"
            ),
            showarrow=False,
            font=dict(
                size=14,
                color="#263238"
            ),
            align="center"
        )

        fig2.update_layout(
            height=360,
            margin=dict(
                l=10,
                r=10,
                t=10,
                b=45
            ),
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.12,
                xanchor="center",
                x=0.5,
                font=dict(size=11)
            ),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)"
        )

        st.plotly_chart(
            fig2,
            use_container_width=True,
            config={
                "displayModeBar": False
            }
        )


    st.divider()


    # ======================================================
    # OVERALL PROGRESS
    # ======================================================

    st.subheader(
        "Overall Cultivation Progress"
    )

    st.progress(
        achievement_pct / 100
    )

    left, right = st.columns(2)

    with left:

        st.success(
            f"✅ Achieved : "
            f"{achieved_hectares:,.2f} Ha "
            f"({achievement_pct:.2f}%)"
        )

    with right:

        st.error(
            f"❌ Pending : "
            f"{pending_hectares:,.2f} Ha "
            f"({pending_pct:.2f}%)"
        )


# ==========================================================
# STUDY & CLUSTER TAB
# ==========================================================

with study_tab:

    st.subheader(
        "🧭 Study & Cluster Cultivation Progress"
    )

    # ------------------------------------------------------
    # Study Progress
    # ------------------------------------------------------

    st.markdown(
        '<div class="section-blue">📊 Study-wise Progress</div>',
        unsafe_allow_html=True
    )

    study_summary = cultivation_summary(
        "Study2"
    )

    styled_study = (
        study_summary.style
        .format(
            {
                "Target_Hectares": "{:.2f}",
                "Achieved_Hectares": "{:.2f}",
                "Pending_Hectares": "{:.2f}",
                "Achieved %": "{:.2f}%",
                "Pending %": "{:.2f}%"
            }
        )
        .map(
            lambda _: (
                "color:green;font-weight:bold;"
            ),
            subset=["Achieved %"]
        )
        .map(
            lambda _: (
                "color:red;font-weight:bold;"
            ),
            subset=["Pending %"]
        )
    )

    st.dataframe(
        styled_study,
        use_container_width=True,
        hide_index=True
    )


    # ------------------------------------------------------
    # Study Cultivation
    # ------------------------------------------------------

    st.markdown(
        '<div class="section-green">🌾 Study-wise Cultivation Summary</div>',
        unsafe_allow_html=True
    )

    study_cp = cultivation_practice_summary(
        "Study2"
    )

    st.dataframe(
        study_cp,
        use_container_width=True,
        hide_index=True
    )


    # ------------------------------------------------------
    # Cluster Cultivation
    # ------------------------------------------------------

    st.markdown(
        '<div class="section-green">🌾 Cluster-wise Cultivation Summary</div>',
        unsafe_allow_html=True
    )

    cluster_cp = cultivation_practice_summary(
        "Cluster"
    )

    st.dataframe(
        cluster_cp,
        use_container_width=True,
        hide_index=True
    )


    # ------------------------------------------------------
    # Cluster Progress
    # ------------------------------------------------------

    st.markdown(
        '<div class="section-blue">📊 Cluster-wise Progress</div>',
        unsafe_allow_html=True
    )

    cluster_summary = cultivation_summary(
        "Cluster"
    )

    styled_cluster = (
        cluster_summary.style
        .format(
            {
                "Target_Hectares": "{:.2f}",
                "Achieved_Hectares": "{:.2f}",
                "Pending_Hectares": "{:.2f}",
                "Achieved %": "{:.2f}%",
                "Pending %": "{:.2f}%"
            }
        )
        .map(
            lambda _: (
                "color:green;font-weight:bold;"
            ),
            subset=["Achieved %"]
        )
        .map(
            lambda _: (
                "color:red;font-weight:bold;"
            ),
            subset=["Pending %"]
        )
    )

    st.dataframe(
        styled_cluster,
        use_container_width=True,
        hide_index=True
    )


# ==========================================================
# FIELD OFFICER TAB
# ==========================================================

with fo_tab:

    st.markdown(
        '<div class="section-purple">👨‍🌾 Field Officer Performance</div>',
        unsafe_allow_html=True
    )

    fo_summary = cultivation_summary(
        "Field Officer"
    )

    styled_fo = (
        fo_summary.style
        .format(
            {
                "Target_Hectares": "{:.2f}",
                "Achieved_Hectares": "{:.2f}",
                "Pending_Hectares": "{:.2f}",
                "Achieved %": "{:.2f}%",
                "Pending %": "{:.2f}%"
            }
        )
        .map(
            lambda _: (
                "color:green;font-weight:bold;"
            ),
            subset=["Achieved %"]
        )
        .map(
            lambda _: (
                "color:red;font-weight:bold;"
            ),
            subset=["Pending %"]
        )
    )

    st.dataframe(
        styled_fo,
        use_container_width=True,
        hide_index=True
    )

    st.divider()

    fig = px.bar(
        fo_summary,
        x="Field Officer",
        y="Achieved %",
        text="Achieved %",
        color="Achieved %",
        color_continuous_scale="Greens",
        title="Field Officer Achievement %"
    )

    fig.update_traces(
        texttemplate="%{text:.2f}%",
        textposition="outside"
    )

    fig.update_layout(
        yaxis_title="Achievement %",
        xaxis_title="Field Officer",
        height=500
    )

    st.plotly_chart(
        fig,
        use_container_width=True
    )

    st.divider()


# ==========================================================
# TIMELINE TAB
# ==========================================================

with timeline_tab:

    st.markdown(
        '<div class="section-orange">📅 Sowing Timeline</div>',
        unsafe_allow_html=True
    )

    timeline_df = filtered_df.copy()

    date_col = (
        "Kharif-26 DSR/Nursery Sowing Date"
    )

    timeline_df = timeline_df[
        timeline_df[date_col].notna()
    ]

    if timeline_df.empty:

        st.warning(
            "No sowing date data available."
        )

    else:

        daily = (
            timeline_df
            .groupby(date_col)
            .agg(
                Plots=("Plot Codes", "nunique"),
                Hectares=(
                    "Kharif-26 Hectares",
                    "sum"
                )
            )
            .reset_index()
            .sort_values(date_col)
        )

        k1, k2 = st.columns(2)

        k1.metric(
            "🌾 Total Sowing Plots",
            f"{daily['Plots'].sum():,}"
        )

        k2.metric(
            "🌱 Total Sown Hectares",
            f"{daily['Hectares'].sum():,.2f}"
        )

        st.divider()

        # Daily plots
        fig = px.line(
            daily,
            x=date_col,
            y="Plots",
            markers=True,
            title="Daily Sowing Progress (Plots)"
        )

        fig.update_layout(
            height=360,
            xaxis_title="Sowing Date",
            yaxis_title="Plots"
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )

        st.divider()

        # Daily hectares
        fig = px.bar(
            daily,
            x=date_col,
            y="Hectares",
            text="Hectares",
            title="Daily Sowing Hectares"
        )

        fig.update_traces(
            texttemplate="%{text:.2f}",
            textposition="outside"
        )

        fig.update_layout(
            height=360,
            xaxis_title="Sowing Date",
            yaxis_title="Hectares"
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )

        st.divider()

        st.dataframe(
            daily,
            use_container_width=True,
            hide_index=True
        )


# ==========================================================
# DATA TAB
# ==========================================================

with data_tab:

    st.markdown(
        '<div class="section-teal">📋 Filtered Data</div>',
        unsafe_allow_html=True
    )

    st.info(
        f"Showing **{len(filtered_df):,}** records"
    )

    st.divider()

    selected_columns = st.multiselect(
        "Select Columns to Display",
        options=filtered_df.columns.tolist(),
        default=filtered_df.columns.tolist()
    )

    display_df = filtered_df[
        selected_columns
    ]

    search = st.text_input(
        "🔍 Search (works across all selected columns)"
    )

    if search:

        mask = (
            display_df
            .astype(str)
            .apply(
                lambda x:
                x.str.contains(
                    search,
                    case=False,
                    na=False
                )
            )
            .any(axis=1)
        )

        display_df = display_df[
            mask
        ]

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        height=600
    )

    st.divider()


# ==========================================================
# FOOTER
# ==========================================================

st.divider()

st.caption(
    "Developed by Jayapal Thoka | Powered by Streamlit"
)
